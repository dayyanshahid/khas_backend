"""One-time helper: convert Schema_Documentation.xlsx -> data/schema_descriptions.json

Usage (from repo root):
    python khas/scripts/excel_to_schema_json.py
    python khas/scripts/excel_to_schema_json.py path/to/file.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path


def cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def convert(xlsx_path: Path, json_path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    data: dict = {"application": "", "conventions": [], "tables": {}}

    for label, value in wb["Overview"].iter_rows(values_only=True):
        lt_raw = "" if label is None else str(label)
        vt = cell(value)
        lt = lt_raw.strip()
        if lt == "Application":
            data["application"] = vt
        elif lt_raw.startswith("  ") and vt:
            data["conventions"].append({"name": lt, "description": vt})

    for row in wb["Tables"].iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        name = cell(row[1])
        data["tables"][name] = {
            "description": cell(row[4]) if len(row) > 4 else "",
            "columns": {},
        }

    for row in wb["Columns"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[2]:
            continue
        tname = cell(row[0])
        cname = cell(row[2])
        desc = cell(row[4]) if len(row) > 4 else ""
        if not desc:
            continue
        data["tables"].setdefault(tname, {"description": "", "columns": {}})
        data["tables"][tname]["columns"][cname] = desc

    wb.close()
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {json_path} ({len(data['tables'])} tables)")


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[2]
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "Schema_Documentation.xlsx"
    out = Path(__file__).resolve().parents[1] / "data" / "schema_descriptions.json"
    convert(xlsx, out)
